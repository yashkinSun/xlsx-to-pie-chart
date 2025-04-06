import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import Wedge
import io
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from openpyxl.drawing.image import Image as XLImage
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class DataVisualizer:
    """
    Класс для визуализации данных о несоответствиях.
    """
    
    def __init__(self):
        """
        Инициализация визуализатора данных.
        """
        # Цвета для отделов
        self.department_colors = {
            "Производство": "royalblue",
            "Офис": "darkorange"
        }
        
        # Размеры графиков
        self.figure_size = (10, 8)
        
    def create_pie_chart(self, analyzer, title="Распределение несоответствий"):
        """
        Создание многоуровневой круговой диаграммы.
        
        Args:
            analyzer (DataAnalyzer): Анализатор данных с результатами анализа
            title (str): Заголовок диаграммы
            
        Returns:
            matplotlib.figure.Figure: Объект фигуры с диаграммой
        """
        # Получаем данные для диаграммы
        labels, sizes, colors, department_labels = analyzer.get_role_data_for_chart()
        
        if not labels:
            raise ValueError("Нет данных для построения диаграммы")
        
        # Создаем фигуру
        fig, ax = plt.subplots(figsize=self.figure_size, subplot_kw=dict(aspect="equal"))
        
        # Создаем внешнее кольцо (роли)
        outer_radius = 1.0
        inner_radius = 0.6
        
        # Вычисляем углы для каждого сегмента
        total = sum(sizes)
        angles = [s/total * 2 * np.pi for s in sizes]
        
        # Начальный угол
        start_angle = 0
        
        # Создаем внешние сегменты (роли)
        outer_wedges = []
        for i, (label, size, color) in enumerate(zip(labels, sizes, colors)):
            end_angle = start_angle + angles[i]
            wedge = Wedge(
                center=(0, 0),
                r=outer_radius,
                theta1=np.degrees(start_angle),
                theta2=np.degrees(end_angle),
                width=outer_radius - inner_radius,
                facecolor=color,
                edgecolor='white',
                linewidth=1,
                alpha=0.7
            )
            ax.add_patch(wedge)
            outer_wedges.append(wedge)
            
            # Добавляем текст для роли
            angle = (start_angle + end_angle) / 2
            x = (outer_radius + inner_radius) / 2 * np.cos(angle)
            y = (outer_radius + inner_radius) / 2 * np.sin(angle)
            
            # Определяем выравнивание текста
            ha = 'left' if x > 0 else 'right'
            va = 'center'
            
            # Добавляем текст
            ax.text(
                x * 1.1, y * 1.1,
                f"{label}\n({size})",
                ha=ha, va=va,
                fontsize=9,
                bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.7)
            )
            
            start_angle = end_angle
        
        # Создаем внутреннее кольцо (отделы)
        # Сначала группируем данные по отделам
        dept_data = {}
        for dept_label, size in zip(department_labels, sizes):
            if dept_label not in dept_data:
                dept_data[dept_label] = 0
            dept_data[dept_label] += size
        
        # Создаем внутренние сегменты (отделы)
        start_angle = 0
        for dept, size in dept_data.items():
            end_angle = start_angle + (size / total * 2 * np.pi)
            wedge = Wedge(
                center=(0, 0),
                r=inner_radius,
                theta1=np.degrees(start_angle),
                theta2=np.degrees(end_angle),
                facecolor=self.department_colors[dept],
                edgecolor='white',
                linewidth=1
            )
            ax.add_patch(wedge)
            
            # Добавляем текст для отдела
            angle = (start_angle + end_angle) / 2
            x = inner_radius * 0.5 * np.cos(angle)
            y = inner_radius * 0.5 * np.sin(angle)
            
            # Добавляем текст
            ax.text(
                x, y,
                f"{dept}\n({size})",
                ha='center', va='center',
                fontsize=10,
                fontweight='bold',
                color='white'
            )
            
            start_angle = end_angle
        
        # Настраиваем оси
        ax.set_xlim(-1.1, 1.1)
        ax.set_ylim(-1.1, 1.1)
        ax.axis('off')
        
        # Добавляем заголовок
        plt.title(title, fontsize=14, pad=20)
        
        # Добавляем легенду для отделов
        legend_elements = [
            plt.Line2D([0], [0], marker='o', color='w', markerfacecolor=color, markersize=10, label=dept)
            for dept, color in self.department_colors.items()
        ]
        ax.legend(handles=legend_elements, loc='lower center', bbox_to_anchor=(0.5, -0.1), ncol=2)
        
        plt.tight_layout()
        
        return fig
    
    def create_comparison_charts(self, current_analyzer, previous_analyzer, 
                                title_current="Текущий месяц", title_previous="Предыдущий месяц"):
        """
        Создание сравнительных круговых диаграмм для текущего и предыдущего периодов.
        
        Args:
            current_analyzer (DataAnalyzer): Анализатор данных с текущими результатами
            previous_analyzer (DataAnalyzer): Анализатор данных с предыдущими результатами
            title_current (str): Заголовок для текущего периода
            title_previous (str): Заголовок для предыдущего периода
            
        Returns:
            matplotlib.figure.Figure: Объект фигуры с диаграммами
        """
        # Создаем фигуру с двумя подграфиками
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8), subplot_kw=dict(aspect="equal"))
        
        # Создаем диаграмму для предыдущего периода
        self._create_pie_subplot(ax1, previous_analyzer, title_previous)
        
        # Создаем диаграмму для текущего периода
        self._create_pie_subplot(ax2, current_analyzer, title_current)
        
        plt.tight_layout()
        
        return fig
    
    def _create_pie_subplot(self, ax, analyzer, title):
        """
        Создание подграфика с круговой диаграммой.
        
        Args:
            ax (matplotlib.axes.Axes): Объект осей для рисования
            analyzer (DataAnalyzer): Анализатор данных с результатами
            title (str): Заголовок диаграммы
        """
        # Получаем данные для диаграммы
        labels, sizes, colors, department_labels = analyzer.get_role_data_for_chart()
        
        if not labels:
            ax.text(0, 0, "Нет данных", ha='center', va='center', fontsize=12)
            ax.axis('off')
            return
        
        # Создаем внешнее кольцо (роли)
        outer_radius = 1.0
        inner_radius = 0.6
        
        # Вычисляем углы для каждого сегмента
        total = sum(sizes)
        angles = [s/total * 2 * np.pi for s in sizes]
        
        # Начальный угол
        start_angle = 0
        
        # Создаем внешние сегменты (роли)
        for i, (label, size, color) in enumerate(zip(labels, sizes, colors)):
            end_angle = start_angle + angles[i]
            wedge = Wedge(
                center=(0, 0),
                r=outer_radius,
                theta1=np.degrees(start_angle),
                theta2=np.degrees(end_angle),
                width=outer_radius - inner_radius,
                facecolor=color,
                edgecolor='white',
                linewidth=1,
                alpha=0.7
            )
            ax.add_patch(wedge)
            
            # Добавляем текст для роли (упрощенный для сравнительных диаграмм)
            angle = (start_angle + end_angle) / 2
            x = (outer_radius + inner_radius) / 2 * np.cos(angle)
            y = (outer_radius + inner_radius) / 2 * np.sin(angle)
            
            # Определяем выравнивание текста
            ha = 'left' if x > 0 else 'right'
            va = 'center'
            
            # Добавляем текст (только название роли)
            ax.text(
                x * 1.1, y * 1.1,
                label,
                ha=ha, va=va,
                fontsize=8
            )
            
            start_angle = end_angle
        
        # Создаем внутреннее кольцо (отделы)
        # Сначала группируем данные по отделам
        dept_data = {}
        for dept_label, size in zip(department_labels, sizes):
            if dept_label not in dept_data:
                dept_data[dept_label] = 0
            dept_data[dept_label] += size
        
        # Создаем внутренние сегменты (отделы)
        start_angle = 0
        for dept, size in dept_data.items():
            end_angle = start_angle + (size / total * 2 * np.pi)
            wedge = Wedge(
                center=(0, 0),
                r=inner_radius,
                theta1=np.degrees(start_angle),
                theta2=np.degrees(end_angle),
                facecolor=self.department_colors[dept],
                edgecolor='white',
                linewidth=1
            )
            ax.add_patch(wedge)
            
            # Добавляем текст для отдела
            angle = (start_angle + end_angle) / 2
            x = inner_radius * 0.5 * np.cos(angle)
            y = inner_radius * 0.5 * np.sin(angle)
            
            # Добавляем текст
            ax.text(
                x, y,
                dept,
                ha='center', va='center',
                fontsize=9,
                fontweight='bold',
                color='white'
            )
            
            start_angle = end_angle
        
        # Настраиваем оси
        ax.set_xlim(-1.1, 1.1)
        ax.set_ylim(-1.1, 1.1)
        ax.axis('off')
        
        # Добавляем заголовок
        ax.set_title(title, fontsize=12, pad=15)
    
    def save_figure_to_file(self, fig, filename):
        """
        Сохранение фигуры в файл.
        
        Args:
            fig (matplotlib.figure.Figure): Объект фигуры
            filename (str): Имя файла для сохранения
            
        Returns:
            str: Путь к сохраненному файлу
        """
        fig.savefig(filename, dpi=300, bbox_inches='tight')
        return filename
    
    def get_figure_as_bytes(self, fig):
        """
        Получение фигуры в виде байтов для вставки в Excel.
        
        Args:
            fig (matplotlib.figure.Figure): Объект фигуры
            
        Returns:
            bytes: Байтовое представление изображения
        """
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        return buf
    
    def create_excel_report(self, data, analyzer, fig, output_file):
        """
        Создание отчета Excel с данными и диаграммой.
        
        Args:
            data (pandas.DataFrame): Исходные данные
            analyzer (DataAnalyzer): Анализатор данных с результатами
            fig (matplotlib.figure.Figure): Объект фигуры с диаграммой
            output_file (str): Имя файла для сохранения
            
        Returns:
            str: Путь к сохраненному файлу
        """
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        
        # Переименовываем первый лист
        ws1 = wb.active
        ws1.title = "Исходные данные"
        
        # Добавляем исходные данные
        for r in dataframe_to_rows(data, index=False, header=True):
            ws1.append(r)
        
        # Создаем лист для сводки
        ws2 = wb.create_sheet(title="Сводка")
        
        # Добавляем заголовок
        ws2.append(["Сводка по ролям"])
        ws2.append([])
        
        # Получаем отсортированный список ролей по трудозатратам
        sorted_roles = analyzer.get_sorted_role_costs()
        
        # Добавляем заголовки столбцов
        ws2.append(["Отдел", "Роль", "Количество", "Трудозатраты (руб.)"])
        
        # Добавляем данные
        for _, row in sorted_roles.iterrows():
            ws2.append([
                row["Отдел"],
                row["Роль"],
                row["Количество"],
                row["Трудозатраты"]
            ])
        
        # Добавляем пустую строку
        ws2.append([])
        
        # Добавляем сводку по отделам
        ws2.append(["Сводка по отделам"])
        ws2.append([])
        
        # Получаем сводку по отделам
        dept_summary = analyzer.get_department_summary()
        
        # Добавляем заголовки столбцов
        ws2.append(["Отдел", "Количество", "Трудозатраты (руб.)"])
        
        # Добавляем данные
        for _, row in dept_summary.iterrows():
            ws2.append([
                row["Отдел"],
                row["Количество"],
                row["Трудозатраты"]
            ])
        
        # Создаем лист для диаграммы
        ws3 = wb.create_sheet(title="Диаграмма")
        
        # Добавляем заголовок
        ws3.append(["Распределение несоответствий"])
        ws3.append([])
        
        # Сохраняем диаграмму во временный файл
        img_buf = self.get_figure_as_bytes(fig)
        
        # Создаем изображение для Excel
        img = XLImage(img_buf)
        
        # Добавляем изображение на лист
        ws3.add_image(img, 'A3')
        
        # Сохраняем книгу
        wb.save(output_file)
        
        return output_file

# Тестирование класса
if __name__ == "__main__":
    from data_loader import DataLoader
    from data_analyzer import DataAnalyzer
    
    # Загружаем данные
    loader = DataLoader()
    loader.load_file("sample_data.xlsx")
    data = loader.get_current_data()
    
    # Анализируем данные
    analyzer = DataAnalyzer()
    analyzer.analyze_data(data)
    
    # Создаем визуализатор
    visualizer = DataVisualizer()
    
    # Создаем круговую диаграмму
    fig = visualizer.create_pie_chart(analyzer)
    
    # Сохраняем диаграмму в файл
    visualizer.save_figure_to_file(fig, "pie_chart.png")
    print("Диаграмма сохранена в файл: pie_chart.png")
    
    # Создаем отчет Excel
    report_file = "report.xlsx"
    visualizer.create_excel_report(data, analyzer, fig, report_file)
    print(f"Отчет сохранен в файл: {report_file}")
    
    # Закрываем фигуру
    plt.close(fig)
